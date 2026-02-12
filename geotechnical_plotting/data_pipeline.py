#!/usr/bin/env python3
"""
SESRO Geotechnical Data Pipeline Module

Comprehensive data loading and processing pipeline for geotechnical plotting.
Handles the complete workflow from CSV loading through unified data object creation.

ARCHITECTURAL OVERVIEW:

Responsibility:
This module provides the complete data pipeline from raw Openground CSVs to
plot-ready unified data objects. It consolidates all data loading, geological
classification, formation grouping, outlier detection, and investigation mapping
into a single cohesive module.

Key Interactions:
- Input: Global parameter mapping CSV + Openground CSVs + Location Details CSV
- Processing: Mapping extraction → CSV loading → Formation enhancement →
  Formation grouping → Outlier detection → Investigation mapping → Unified objects
- Output: plotly_unified_data + manually_modified_data dictionaries
- Zero CONFIG dependencies (all primitives injected via function parameters)

Navigation Guide:
Use VS Code outline (Ctrl+Shift+O) to navigate between sections:
- Section 1: Parameter Mapping Extraction
- Section 2: CSV Loading & Column Standardization
- Section 3: Geological Formation Processing
- Section 4: Formation Grouping & Data Organization
- Section 5: Outlier Detection
- Section 6: Investigation Mapping & Unified Data Objects
- Section 7: Excel Export
"""

# Standard library imports
import logging
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Optional, Tuple, Any

# Third-party imports
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Logging setup
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# 📊 SECTION 1: PARAMETER MAPPING EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════


def extract_parameter_mappings(
    parameter_name: str,
    mapping_csv_path: str,
) -> pd.DataFrame:
    """
    Extract CSV-to-column mappings for specified parameter from global mapping file.

    This function parses Parameter_CSV_Mapping.csv to identify
    all CSV files and columns containing data for the target parameter. Results are sorted
    by priority_rank (lower is better) and quality_score (higher is better).

    Args:
        parameter_name: Target parameter matching 'parameter' column in mapping CSV
                       (e.g., "UndrainedShearStrength", "MoistureContent")
        mapping_csv_path: Path to Parameter_CSV_Mapping.csv

    Returns:
        DataFrame with columns: [csv_file, column_name, priority_rank, quality_score]
        Sorted by priority_rank ascending, quality_score descending
        Empty DataFrame if parameter not found

    Raises:
        FileNotFoundError: If mapping_csv_path does not exist
        KeyError: If required columns missing from mapping CSV
    """
    logger.info(f"🚀 PHASE 1: Extracting parameter mappings for '{parameter_name}'")

    # === VALIDATION SECTION ===
    mapping_path = Path(mapping_csv_path)
    if not mapping_path.exists():
        logger.error(f"❌ Mapping file not found: {mapping_csv_path}")
        raise FileNotFoundError(f"Mapping CSV not found: {mapping_csv_path}")

    # === DATA LOADING SECTION ===
    try:
        mapping_df = pd.read_csv(mapping_csv_path)
        logger.info(
            f"📊 Loaded mapping CSV: {len(mapping_df)} total parameter mappings"
        )
    except Exception as e:
        logger.error(f"❌ Failed to read mapping CSV: {e}")
        raise

    # === COLUMN VALIDATION SECTION ===
    required_cols = [
        "parameter",
        "csv_file",
        "column_name",
        "priority_rank",
        "quality_score",
    ]
    missing_cols = [col for col in required_cols if col not in mapping_df.columns]
    if missing_cols:
        logger.error(f"❌ Missing required columns in mapping CSV: {missing_cols}")
        raise KeyError(f"Mapping CSV missing required columns: {missing_cols}")

    # === PARAMETER FILTERING SECTION ===
    param_mappings = mapping_df[mapping_df["parameter"] == parameter_name].copy()

    if param_mappings.empty:
        logger.warning(f"⚠️ No mappings found for parameter '{parameter_name}'")
        logger.warning(
            f"Available parameters: {sorted(mapping_df['parameter'].unique())}"
        )
        return pd.DataFrame(columns=required_cols)

    # === SORTING SECTION ===
    param_mappings = param_mappings.sort_values(
        by=["priority_rank", "quality_score"], ascending=[True, False]
    ).reset_index(drop=True)

    # === LOGGING SECTION ===
    logger.info(f"✅ Found {len(param_mappings)} CSV sources for '{parameter_name}':")
    for idx, row in param_mappings.iterrows():
        logger.info(
            f"  {idx+1}. {row['csv_file']:<45} → {row['column_name']:<30} "
            f"(Priority: {row['priority_rank']}, Quality: {row['quality_score']})"
        )

    return param_mappings[required_cols]


# ═══════════════════════════════════════════════════════════════════════════
# 💾 SECTION 2: CSV LOADING & COLUMN STANDARDIZATION
# ═══════════════════════════════════════════════════════════════════════════


def filter_consolidation_loading_phase(
    df: pd.DataFrame,
    location_column: str,
    cell_pressure_column: str,
    column_variations: Dict[str, List[str]],
) -> pd.DataFrame:
    """
    Remove unloading phase data after maximum cell pressure for each location AND depth.

    For consolidation tests, data often includes both loading and unloading phases.
    This function keeps only the loading phase data (up to and including the maximum
    cell pressure) for each unique Location ID + Top Depth combination.

    Args:
        df: Consolidation test data with location and cell pressure columns
        location_column: Name of column containing Location ID
        cell_pressure_column: Name of column containing cell pressure values
        column_variations: Column name variations for finding depth column

    Returns:
        Filtered dataframe containing only loading phase data

    Algorithm:
        1. Group by Location ID + Top Depth
        2. For each group:
           a. Sort by Increment No (test sequence)
           b. Find index of first maximum cell pressure value
           c. Keep only rows up to and including maximum
        3. Combine filtered groups
    """
    # Find depth column
    depth_col = None
    for col_name in column_variations.get("top_depth", ["Top Depth"]):
        if col_name in df.columns:
            depth_col = col_name
            break

    if depth_col is None:
        logger.warning(
            "⚠️ Consolidation filtering: No depth column found - skipping filter"
        )
        return df

    if location_column not in df.columns:
        logger.warning(
            f"⚠️ Consolidation filtering: Location column '{location_column}' not found"
        )
        return df

    if cell_pressure_column not in df.columns:
        logger.warning(
            f"⚠️ Consolidation filtering: Cell pressure column '{cell_pressure_column}' not found"
        )
        return df

    # === FILTER EACH LOCATION + DEPTH GROUP ===
    filtered_groups = []
    group_columns = [location_column, depth_col]

    logger.info(
        f"🔍 Consolidation filtering: Processing {len(df)} rows by Location+Depth"
    )

    for (location_id, depth), group_df in df.groupby(group_columns):
        # Sort by Increment No to get correct test sequence
        if "Increment No" in group_df.columns:
            group_df = group_df.sort_values("Increment No").reset_index(drop=True)
        else:
            group_df = group_df.reset_index(drop=True)

        # Check for NaN values in cell pressure
        cell_pressure_values = group_df[cell_pressure_column]
        nan_count = cell_pressure_values.isna().sum()

        if nan_count == len(cell_pressure_values):
            logger.debug(
                f"  ⚠️ {location_id} @ {depth}m - all cell pressure values NaN, skipping"
            )
            continue

        # Find index of maximum cell pressure (first occurrence)
        max_idx = group_df[cell_pressure_column].idxmax()

        if pd.isna(max_idx):
            logger.debug(f"  ⚠️ {location_id} @ {depth}m - no valid max found, skipping")
            continue

        # Keep only rows up to and including maximum
        loading_phase = group_df.loc[:max_idx].copy()

        logger.debug(
            f"  ✓ {location_id} @ {depth}m: kept {len(loading_phase)}/{len(group_df)} rows"
        )

        filtered_groups.append(loading_phase)

    if not filtered_groups:
        logger.warning("⚠️ Consolidation filtering: No valid groups found")
        return df

    result_df = pd.concat(filtered_groups, ignore_index=True)
    removed_count = len(df) - len(result_df)
    logger.info(
        f"✅ Consolidation filtering: Removed {removed_count} unloading phase rows"
    )

    return result_df


def find_column_in_dataframe(
    df: pd.DataFrame,
    column_type: str,
    column_variations: Dict[str, List[str]],
    csv_name: Optional[str] = None,
) -> Optional[str]:
    """
    Find the correct column name in a DataFrame by checking all possible variations.

    Args:
        df: pandas DataFrame to search for columns
        column_type: String key from column_variations (e.g., 'geology_code2')
        column_variations: Dictionary mapping column types to lists of possible names
        csv_name: Optional CSV name for debugging purposes

    Returns:
        str or None: The found column name, or None if not found
    """
    if column_type not in column_variations:
        logger.warning(f"⚠️ Unknown column type '{column_type}' requested")
        return None

    available_columns = df.columns.tolist()
    possible_names = column_variations[column_type]

    # Check each possible name (exact match first)
    for name in possible_names:
        if name in available_columns:
            logger.debug(
                f"✅ Found column '{name}' for type '{column_type}' in {csv_name or 'DataFrame'}"
            )
            return name

    # If no exact match, try case-insensitive matching
    for name in possible_names:
        for col in available_columns:
            if name.lower() == col.lower():
                logger.debug(
                    f"✅ Found column '{col}' (case-insensitive) for type '{column_type}' in {csv_name or 'DataFrame'}"
                )
                return col

    logger.warning(
        f"❌ Column type '{column_type}' not found in {csv_name or 'DataFrame'}. "
        f"Available columns: {', '.join(available_columns[:5])}{'...' if len(available_columns) > 5 else ''}"
    )
    return None


def standardize_location_columns(
    df: pd.DataFrame,
    csv_name: str,
    column_variations: Dict[str, List[str]],
) -> pd.DataFrame:
    """
    Standardize location and depth column names via safe copy operation.

    Args:
        df: DataFrame to standardize
        csv_name: CSV filename for logging
        column_variations: Dictionary of column name variations

    Returns:
        DataFrame with standardized 'Location ID' and 'Top Depth' columns
    """
    df_copy = df.copy()

    # Standardize location column
    location_col = find_column_in_dataframe(
        df_copy, "location_id", column_variations, csv_name
    )
    if location_col is not None:
        df_copy["Location ID"] = df_copy[location_col]
        logger.debug(f"✅ {csv_name}: Mapped '{location_col}' to 'Location ID'")
    elif "Location ID" not in df_copy.columns:
        logger.warning(
            f"WARNING {csv_name}: No location column found - will skip location-based analysis"
        )

    # Standardize depth column
    depth_col = find_column_in_dataframe(
        df_copy, "top_depth", column_variations, csv_name
    )
    if depth_col is not None:
        df_copy["Top Depth"] = df_copy[depth_col]
        logger.debug(f"✅ {csv_name}: Mapped '{depth_col}' to 'Top Depth'")
    elif "Top Depth" not in df_copy.columns:
        logger.warning(
            f"WARNING {csv_name}: No depth column found - will use default values"
        )

    return df_copy


def filter_location_ids(
    df: pd.DataFrame,
    csv_name: str,
    enabled: bool,
    exclude_prefixes: List[str],
    case_sensitive: bool,
    column_variations: Dict[str, List[str]],
) -> pd.DataFrame:
    """
    Filter out rows with excluded location ID prefixes.

    Args:
        df: DataFrame to filter
        csv_name: CSV filename for logging
        enabled: Whether filtering is enabled
        exclude_prefixes: List of prefixes to exclude (e.g., ["BP", "CCT"])
        case_sensitive: Whether to use case-sensitive matching
        column_variations: Dictionary of column name variations

    Returns:
        Filtered DataFrame
    """
    if not enabled:
        return df

    df_copy = df.copy()
    initial_count = len(df_copy)

    location_col = find_column_in_dataframe(
        df_copy, "location_id", column_variations, csv_name
    )

    if location_col is None:
        logger.debug(f"INFO {csv_name}: No location column found for filtering")
        return df_copy

    filter_mask = pd.Series([True] * len(df_copy), index=df_copy.index)

    for prefix in exclude_prefixes:
        if case_sensitive:
            prefix_filter = (
                df_copy[location_col].astype(str).str.startswith(prefix, na=False)
            )
        else:
            prefix_filter = (
                df_copy[location_col]
                .astype(str)
                .str.upper()
                .str.startswith(prefix.upper(), na=False)
            )
        filter_mask = filter_mask & ~prefix_filter

    df_copy = df_copy[filter_mask]

    filtered_count = len(df_copy)
    excluded_count = initial_count - filtered_count

    if excluded_count > 0:
        logger.info(
            f"🚫 {csv_name}: Filtered out {excluded_count} rows with excluded location prefixes ({exclude_prefixes})"
        )

    return df_copy


def apply_formation_specific_filters(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    formation_specific_filters: Dict[str, List[str]],
    geology_code_descriptions: Dict[str, str],
    case_sensitive: bool = False,
) -> int:
    """
    Apply formation-specific location ID filters to grouped data.

    Filters out specific location prefixes from specific formations only.
    For example, exclude OS01/OS02 from KC_W and KC_UW formations only.

    Args:
        formation_groups: Dict of formation_name -> {csv_name: DataFrame}
            MODIFIED IN-PLACE
        formation_specific_filters: Dict of prefix -> [formation_codes]
            e.g., {"OS01": ["KC_W", "KC_UW"]} excludes OS01 from KC formations
        geology_code_descriptions: Dict of formation_code -> description
            e.g., {"KC_W": "Kimmeridge Clay Formation (Weathered)"}
        case_sensitive: Whether prefix matching is case-sensitive

    Returns:
        Total number of rows filtered out
    """
    if not formation_specific_filters:
        return 0

    # Build mapping: formation_description (upper) -> formation_code (upper)
    description_to_code = {}
    for code, description in geology_code_descriptions.items():
        description_to_code[description.upper()] = code.upper()

    total_filtered = 0

    for formation_name, csv_data in formation_groups.items():
        # Get formation code from name
        formation_code = description_to_code.get(formation_name.upper(), "")

        # Find which prefixes apply to this formation
        applicable_prefixes = []
        for prefix, formation_codes in formation_specific_filters.items():
            formation_codes_upper = [fc.upper() for fc in formation_codes]
            if formation_code in formation_codes_upper:
                applicable_prefixes.append(prefix)

        if not applicable_prefixes:
            continue

        # Apply filtering to each CSV in this formation
        for csv_name, df in csv_data.items():
            if df.empty:
                continue

            # Find location column
            location_col = None
            for col in df.columns:
                if col.upper() in ["LOCATION ID", "LOCATIONID", "LOCATION_ID"]:
                    location_col = col
                    break

            if location_col is None:
                continue

            initial_count = len(df)
            filter_mask = pd.Series([True] * len(df), index=df.index)

            for prefix in applicable_prefixes:
                if case_sensitive:
                    prefix_filter = (
                        df[location_col].astype(str).str.startswith(prefix, na=False)
                    )
                else:
                    prefix_filter = (
                        df[location_col]
                        .astype(str)
                        .str.upper()
                        .str.startswith(prefix.upper(), na=False)
                    )
                filter_mask = filter_mask & ~prefix_filter

            filtered_df = df[filter_mask]
            excluded_count = initial_count - len(filtered_df)

            if excluded_count > 0:
                csv_data[csv_name] = filtered_df
                total_filtered += excluded_count
                logger.info(
                    f"🚫 {csv_name} [{formation_name}]: Filtered {excluded_count} rows "
                    f"with formation-specific prefixes {applicable_prefixes}"
                )

    return total_filtered


# ═══════════════════════════════════════════════════════════════════════════
# 🌍 SECTION 3: GEOLOGICAL FORMATION PROCESSING
# ═══════════════════════════════════════════════════════════════════════════


def _get_geology_column_names(
    df: pd.DataFrame,
    csv_name: str,
    column_variations: Dict[str, List[str]],
) -> Tuple[Optional[str], Optional[str]]:
    """Get the geology column names by finding them in the DataFrame."""
    geology_desc_col = find_column_in_dataframe(
        df, "geology_code_description", column_variations, csv_name
    )
    geology_code_col = find_column_in_dataframe(
        df, "geology_code", column_variations, csv_name
    )
    return geology_desc_col, geology_code_col


def _find_geology_description_column(
    df_copy: pd.DataFrame, csv_name: str, geology_desc_col: Optional[str]
) -> bool:
    """Find and map geology description column to standard name."""
    if geology_desc_col and geology_desc_col in df_copy.columns:
        logger.debug(f"INFO {csv_name}: Found {geology_desc_col}")
        df_copy["Geological_Strata"] = df_copy[geology_desc_col]
        return True
    elif "Geology Code Description" in df_copy.columns:
        logger.debug(f"INFO {csv_name}: Found standard 'Geology Code Description'")
        df_copy["Geological_Strata"] = df_copy["Geology Code Description"]
        return True
    return False


def _create_description_from_code(
    df_copy: pd.DataFrame,
    csv_name: str,
    geology_code_col: Optional[str],
    geology_code_descriptions: Dict[str, str],
) -> None:
    """Create geology description from geology code using mapping."""
    if geology_code_col and geology_code_col in df_copy.columns:
        logger.debug(
            f"CREATING DESCRIPTION: {csv_name}: Creating description from {geology_code_col}"
        )
        df_copy["Geological_Strata"] = df_copy[geology_code_col].map(
            geology_code_descriptions
        )
        df_copy["Geological_Strata"].fillna(df_copy[geology_code_col], inplace=True)
    elif "Geology Code" in df_copy.columns:
        logger.debug(
            f"INFO {csv_name}: Creating description from standard 'Geology Code'"
        )
        df_copy["Geological_Strata"] = df_copy["Geology Code"].map(
            geology_code_descriptions
        )
        df_copy["Geological_Strata"].fillna(df_copy["Geology Code"], inplace=True)
    else:
        logger.error(
            f"❌ GEOLOGY PROCESSING FAILED: {csv_name} - No geology code column found"
        )
        df_copy["Geological_Strata"] = "Unknown Formation"


def classify_weathering_state(
    geology_code: str,
    geology_code2: Optional[str],
    geology_code2_desc: Optional[str],
    weathered_suffix: str,
    unweathered_suffix: str,
    fallback_classification: str,
) -> str:
    """Classify weathering state based on Geology Code2 prefix and suffix patterns."""
    if not geology_code2 or pd.isna(geology_code2):
        return fallback_classification

    geology_code2_clean = str(geology_code2).strip()
    expected_prefix = f"{geology_code}_"

    if not geology_code2_clean.startswith(expected_prefix):
        return fallback_classification

    if geology_code2_clean.endswith(weathered_suffix):
        return "weathered"

    if geology_code2_clean.endswith(unweathered_suffix):
        return "unweathered"

    return fallback_classification


def classify_rtd_subformation(
    geology_code: str,
    geology_code2: Optional[str],
    geology_code2_desc: Optional[str],
    rtd_mapping: Dict[str, str],
    fallback_classification: str,
) -> str:
    """Classify RTD sub-formation based on Geology Code2 patterns."""
    if geology_code != "RTD":
        return fallback_classification

    if not geology_code2 or pd.isna(geology_code2):
        return fallback_classification

    geology_code2_clean = str(geology_code2).strip()

    if geology_code2_clean in rtd_mapping:
        return rtd_mapping[geology_code2_clean]

    return fallback_classification


def create_enhanced_formation_code(
    geology_code: str,
    classification_state: str,
    target_formations: List[str],
    formation_type: str = "weathering",
) -> str:
    """Create enhanced formation code with weathering or RTD sub-formation designation."""
    if geology_code not in target_formations:
        return geology_code

    # Handle RTD formations
    if formation_type == "rtd" and geology_code == "RTD":
        if classification_state == "rtd_1":
            return "RTD_1"
        elif classification_state == "rtd_2":
            return "RTD_2"
        elif classification_state == "rtd_3":
            return "RTD_3"
        elif classification_state == "undefined":
            return "RTD_Undefined"
        else:
            return geology_code

    # Handle weathering formations (GF, KC)
    elif formation_type == "weathering" and geology_code in ["GF", "KC"]:
        if classification_state == "weathered":
            return f"{geology_code}_W"
        elif classification_state == "unweathered":
            return f"{geology_code}_UW"
        elif classification_state == "undefined":
            return geology_code
        else:
            return geology_code

    return geology_code


def _get_geology_code2_column_names(
    df: pd.DataFrame,
    csv_name: str,
    column_variations: Dict[str, List[str]],
) -> Tuple[Optional[str], Optional[str]]:
    """Get the correct Geology Code2 column names."""
    geology_code2_col = find_column_in_dataframe(
        df, "geology_code2", column_variations, csv_name
    )
    geology_code2_desc_col = find_column_in_dataframe(
        df, "geology_code2_description", column_variations, csv_name
    )
    return geology_code2_col, geology_code2_desc_col


def add_strata_description(
    df: pd.DataFrame,
    csv_name: str,
    geology_code_descriptions: Dict[str, str],
    column_variations: Dict[str, List[str]],
) -> pd.DataFrame:
    """Add Geology Code Description column if missing and standardize strata column."""
    df_copy = df.copy()
    geology_desc_col, geology_code_col = _get_geology_column_names(
        df_copy, csv_name, column_variations
    )

    geology_desc_found = _find_geology_description_column(
        df_copy, csv_name, geology_desc_col
    )

    if not geology_desc_found and geology_code_col is not None:
        _create_description_from_code(
            df_copy, csv_name, geology_code_col, geology_code_descriptions
        )

    # Handle missing strata values
    missing_strata = df_copy["Geological_Strata"].isna().sum()
    if missing_strata > 0:
        logger.warning(
            f"WARNING {csv_name}: {missing_strata} rows with missing strata description"
        )

    return df_copy


def _apply_formation_splitting(
    df: pd.DataFrame,
    csv_name: str,
    geology_code_col: str,
    geology_code2_col: str,
    geology_code2_desc_col: Optional[str],
    target_formations: List[str],
    weathering_config: Dict[str, str],
    geology_code_descriptions: Dict[str, str],
    rtd_target_formations: Optional[List[str]] = None,
    rtd_classification_config: Optional[Dict[str, Any]] = None,
) -> pd.DataFrame:
    """
    Apply formation splitting logic for both weathering and RTD classifications.

    PERFORMANCE: Vectorized approach using np.select is 10-30× faster than iterrows.
    Uses vectorized string operations (.str.endswith, .isin) and numpy conditionals.
    """
    df_copy = df.copy()

    weathered_suffix = weathering_config["weathered_suffix"]
    unweathered_suffix = weathering_config["unweathered_suffix"]
    fallback_classification = weathering_config["fallback_classification"]

    # === VECTORIZED WEATHERING CLASSIFICATION ===
    # Get geology code columns as series
    geology_codes = df_copy[geology_code_col].astype(str)
    geology_codes2 = df_copy[geology_code2_col].astype(str).str.strip()

    # Mask for target formations (GF, KC, etc.)
    target_mask = geology_codes.isin(target_formations)

    # Build expected prefixes for each row (e.g., "GF_", "KC_")
    expected_prefixes = geology_codes + "_"

    # Check if geology_code2 starts with expected prefix
    prefix_match = pd.Series(
        [
            str(g2).startswith(str(ep)) if pd.notna(g2) else False
            for g2, ep in zip(geology_codes2, expected_prefixes)
        ],
        index=df_copy.index,
    )

    # Vectorized suffix checking (only for rows with matching prefix)
    ends_with_weathered = geology_codes2.str.endswith(weathered_suffix, na=False)
    ends_with_unweathered = geology_codes2.str.endswith(unweathered_suffix, na=False)

    # Classification conditions
    is_weathered = target_mask & prefix_match & ends_with_weathered
    is_unweathered = target_mask & prefix_match & ends_with_unweathered
    is_undefined = target_mask & ~(is_weathered | is_unweathered)

    # Count for logging
    weathered_count = is_weathered.sum()
    unweathered_count = is_unweathered.sum()
    undefined_count = is_undefined.sum()

    # Apply weathered classifications vectorized
    for formation in target_formations:
        formation_mask = geology_codes == formation

        # Weathered
        w_mask = formation_mask & is_weathered
        enhanced_code_w = f"{formation}_W"
        if enhanced_code_w in geology_code_descriptions:
            df_copy.loc[w_mask, "Geological_Strata"] = geology_code_descriptions[
                enhanced_code_w
            ]

        # Unweathered
        uw_mask = formation_mask & is_unweathered
        enhanced_code_uw = f"{formation}_UW"
        if enhanced_code_uw in geology_code_descriptions:
            df_copy.loc[uw_mask, "Geological_Strata"] = geology_code_descriptions[
                enhanced_code_uw
            ]

        # Undefined
        undef_mask = formation_mask & is_undefined
        if formation in geology_code_descriptions:
            base_description = geology_code_descriptions[formation]
            df_copy.loc[undef_mask, "Geological_Strata"] = (
                f"{base_description} (Undefined Classification)"
            )

    # === VECTORIZED RTD CLASSIFICATION ===
    rtd_1_count = 0
    rtd_2_count = 0
    rtd_3_count = 0
    rtd_undefined_count = 0

    if rtd_target_formations and rtd_classification_config:
        rtd_mapping = rtd_classification_config["mapping"]

        # Mask for RTD formations
        rtd_mask = geology_codes.isin(rtd_target_formations)

        if rtd_mask.any():
            # Map geology_code2 to RTD subformation using the mapping dict
            rtd_subformations = geology_codes2.map(rtd_mapping)

            # Classify based on mapped values
            is_rtd_1 = rtd_mask & (rtd_subformations == "rtd_1")
            is_rtd_2 = rtd_mask & (rtd_subformations == "rtd_2")
            is_rtd_3 = rtd_mask & (rtd_subformations == "rtd_3")
            is_rtd_undefined = rtd_mask & rtd_subformations.isna()

            rtd_1_count = is_rtd_1.sum()
            rtd_2_count = is_rtd_2.sum()
            rtd_3_count = is_rtd_3.sum()
            rtd_undefined_count = is_rtd_undefined.sum()

            # Apply RTD descriptions
            if "RTD_1" in geology_code_descriptions:
                df_copy.loc[is_rtd_1, "Geological_Strata"] = geology_code_descriptions[
                    "RTD_1"
                ]
            if "RTD_2" in geology_code_descriptions:
                df_copy.loc[is_rtd_2, "Geological_Strata"] = geology_code_descriptions[
                    "RTD_2"
                ]
            if "RTD_3" in geology_code_descriptions:
                df_copy.loc[is_rtd_3, "Geological_Strata"] = geology_code_descriptions[
                    "RTD_3"
                ]
            if "RTD_Undefined" in geology_code_descriptions:
                df_copy.loc[is_rtd_undefined, "Geological_Strata"] = (
                    geology_code_descriptions["RTD_Undefined"]
                )

    # Calculate splitting_applied for logging
    splitting_applied = (
        weathered_count
        + unweathered_count
        + undefined_count
        + rtd_1_count
        + rtd_2_count
        + rtd_3_count
        + rtd_undefined_count
    )

    # Log results
    if splitting_applied > 0:
        if weathered_count + unweathered_count + undefined_count > 0:
            logger.info(
                f"✅ {csv_name}: Weathering classification applied to {weathered_count + unweathered_count + undefined_count} rows "
                f"(W:{weathered_count}, UW:{unweathered_count}, Undefined:{undefined_count})"
            )
        if rtd_1_count + rtd_2_count + rtd_3_count + rtd_undefined_count > 0:
            logger.info(
                f"✅ {csv_name}: RTD classification applied to {rtd_1_count + rtd_2_count + rtd_3_count + rtd_undefined_count} rows "
                f"(RTD_1:{rtd_1_count}, RTD_2:{rtd_2_count}, RTD_3:{rtd_3_count}, Undefined:{rtd_undefined_count})"
            )

    return df_copy


def create_enhanced_geological_strata_column(
    df: pd.DataFrame,
    csv_name: str,
    geology_code_descriptions: Dict[str, str],
    column_variations: Dict[str, List[str]],
    formation_splitting_config: Dict[str, Any],
) -> pd.DataFrame:
    """Enhanced formation mapping with weathered/unweathered splitting capability."""
    # Start with standard geological enhancement
    df_enhanced = add_strata_description(
        df,
        csv_name,
        geology_code_descriptions,
        column_variations,
    )

    if not formation_splitting_config.get("enabled", False):
        logger.debug(
            f"INFO {csv_name}: Formation splitting disabled, using standard processing"
        )
        return df_enhanced

    target_formations = formation_splitting_config["target_formations"]
    weathering_config = formation_splitting_config["weathering_classification"]

    # Extract RTD configuration if enabled
    rtd_target_formations = None
    rtd_classification_config = None
    if formation_splitting_config.get("rtd_classification", {}).get("enabled", False):
        rtd_target_formations = formation_splitting_config["rtd_classification"][
            "target_formations"
        ]
        rtd_classification_config = formation_splitting_config["rtd_classification"]

    # Get Geology Code2 column names
    geology_code2_col, geology_code2_desc_col = _get_geology_code2_column_names(
        df_enhanced, csv_name, column_variations
    )

    if geology_code2_col is None or geology_code2_col not in df_enhanced.columns:
        logger.debug(
            f"INFO {csv_name}: No Geology Code 2 column found, using standard strata"
        )
        return df_enhanced

    logger.info(f"🔄 {csv_name}: Applying formation splitting for {target_formations}")

    # Get geology code column
    geology_code_col = find_column_in_dataframe(
        df_enhanced, "geology_code", column_variations, csv_name
    )
    if geology_code_col is None or geology_code_col not in df_enhanced.columns:
        logger.warning(f"WARNING {csv_name}: No geology code column found")
        return df_enhanced

    df_enhanced = _apply_formation_splitting(
        df_enhanced,
        csv_name,
        geology_code_col,
        geology_code2_col,
        geology_code2_desc_col,
        target_formations,
        weathering_config,
        geology_code_descriptions,
        rtd_target_formations,
        rtd_classification_config,
    )

    return df_enhanced


def load_csv_data(
    csv_files: List[Path],
    geology_code_descriptions: Dict[str, str],
    column_variations: Dict[str, List[str]],
    formation_splitting_config: Dict[str, Any],
    filter_enabled: bool,
    exclude_prefixes: List[str],
    case_sensitive: bool,
    consolidation_filtering_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, pd.DataFrame]:
    """
    Load and standardize all CSV files with geological enhancement.

    Args:
        csv_files: List of Path objects to CSV files
        geology_code_descriptions: Mapping of geology codes to descriptions
        column_variations: Dictionary of column name variations
        formation_splitting_config: Configuration for formation splitting
        filter_enabled: Whether location filtering is enabled
        exclude_prefixes: List of location prefixes to exclude
        case_sensitive: Whether location filtering is case-sensitive
        consolidation_filtering_config: Config for consolidation loading phase filter (optional)

    Returns:
        Dictionary mapping CSV filenames to enhanced DataFrames
    """
    logger.info(f"📊 Loading {len(csv_files)} CSV files...")
    csv_data_dict = {}

    for csv_file in csv_files:
        if isinstance(csv_file, str):
            csv_path = Path(csv_file)
            csv_name = Path(csv_file).name
        else:
            csv_path = csv_file
            csv_name = csv_file.name

        logger.debug(f"LOADING Loading {csv_name}...")

        try:
            df = pd.read_csv(csv_path, encoding="utf-8")

            if df.empty:
                logger.warning(f"WARNING {csv_name}: CSV file is empty, skipping.")
                continue

            # Defensive data handling
            df.columns = df.columns.str.strip()
            logger.debug(f"  STATS Loaded {len(df)} rows, {len(df.columns)} columns")

            # Add strata description with enhanced formation splitting
            df_with_strata = create_enhanced_geological_strata_column(
                df,
                csv_name,
                geology_code_descriptions,
                column_variations,
                formation_splitting_config,
            )
            if df_with_strata is None:
                continue

            # Standardize location columns
            standardized_df = standardize_location_columns(
                df_with_strata, csv_name, column_variations
            )
            if standardized_df is None:
                continue

            # Apply location ID filtering
            filtered_df = filter_location_ids(
                standardized_df,
                csv_name,
                filter_enabled,
                exclude_prefixes,
                case_sensitive,
                column_variations,
            )
            if filtered_df is None:
                logger.error(
                    f"❌ CSV PROCESSING FAILED: {csv_name} - Failed during filtering"
                )
                continue

            # === CONSOLIDATION LOADING PHASE FILTER ===
            if (
                consolidation_filtering_config
                and consolidation_filtering_config.get("enabled", False)
                and csv_name in consolidation_filtering_config.get("apply_to_csvs", [])
            ):
                # Find cell pressure and location columns
                cell_pressure_col = find_column_in_dataframe(
                    filtered_df, "cell_pressure", column_variations, csv_name
                )
                location_col = find_column_in_dataframe(
                    filtered_df, "location_id", column_variations, csv_name
                )

                if cell_pressure_col and location_col:
                    logger.info(
                        f"🔍 Applying consolidation loading phase filter to {csv_name}..."
                    )
                    filtered_df = filter_consolidation_loading_phase(
                        df=filtered_df,
                        location_column=location_col,
                        cell_pressure_column=cell_pressure_col,
                        column_variations=column_variations,
                    )
                else:
                    if not cell_pressure_col:
                        logger.warning(
                            f"⚠️ {csv_name}: Cell pressure column not found - skipping consolidation filtering"
                        )
                    if not location_col:
                        logger.warning(
                            f"⚠️ {csv_name}: Location column not found - skipping consolidation filtering"
                        )

            csv_data_dict[csv_name] = filtered_df
            logger.info(f"✅ {csv_name}: Successfully processed")

        except Exception as e:
            logger.error(f"❌ CSV LOAD FAILED: {csv_name} - {str(e)}")

    logger.info(f"✅ Successfully loaded {len(csv_data_dict)} CSV files")
    return csv_data_dict


def apply_row_level_fallback(
    csv_data_dict: Dict[str, pd.DataFrame],
    parameter_mappings: pd.DataFrame,
    fallback_parameters: List[str],
) -> None:
    """
    Apply row-level fallback for parameters with multiple column sources.

    For each parameter that has multiple column mappings (e.g., primary + backup),
    this function creates a unified column that uses the primary column value
    when available, and falls back to backup column(s) when primary is NaN.

    This enables data recovery from backup sources (e.g., 'Reported Coefficient
    of Consolidation' when 'Coefficient of consolidation Root Time' is missing).

    Args:
        csv_data_dict: Dictionary of loaded CSV dataframes (MODIFIED IN-PLACE)
        parameter_mappings: DataFrame with columns [parameter, csv_file, column_name,
                           priority_rank, quality_score]
        fallback_parameters: List of parameter names to apply fallback to

    Side Effects:
        Modifies csv_data_dict in-place by adding unified parameter columns
        named '{parameter_name}_unified' to each CSV dataframe.
    """
    if not fallback_parameters:
        logger.debug("⏭️ No parameters configured for fallback - skipping")
        return

    logger.info("🔄 Applying row-level fallback for parameters with multiple sources")

    # Process each parameter that needs fallback
    for parameter_name in fallback_parameters:
        param_mappings = parameter_mappings[
            parameter_mappings["parameter"] == parameter_name
        ]

        if param_mappings.empty:
            logger.warning(
                f"⚠️ Parameter '{parameter_name}' not found in mappings - skipping fallback"
            )
            continue

        # Group mappings by CSV file
        for csv_name in param_mappings["csv_file"].unique():
            if csv_name not in csv_data_dict:
                continue

            df = csv_data_dict[csv_name]
            csv_param_mappings = param_mappings[
                param_mappings["csv_file"] == csv_name
            ].sort_values("priority_rank")

            # Get column names in priority order
            column_names = csv_param_mappings["column_name"].tolist()

            # Filter to columns that actually exist in the dataframe
            existing_columns = [col for col in column_names if col in df.columns]

            if len(existing_columns) == 0:
                logger.warning(
                    f"⚠️ {csv_name}: No columns found for '{parameter_name}' - skipping"
                )
                continue

            if len(existing_columns) == 1:
                # Only one column exists - no fallback needed, just alias it
                unified_col_name = f"{parameter_name}_unified"
                df[unified_col_name] = pd.to_numeric(
                    df[existing_columns[0]], errors="coerce"
                )
                logger.debug(
                    f"  {csv_name}: '{parameter_name}' - single source (aliased '{existing_columns[0]}')"
                )
                continue

            # Multiple columns exist - apply fallback logic
            unified_col_name = f"{parameter_name}_unified"

            # Start with the primary column (priority rank 1)
            df[unified_col_name] = pd.to_numeric(
                df[existing_columns[0]], errors="coerce"
            )

            # Track fallback statistics
            initial_null_count = df[unified_col_name].isna().sum()
            fallback_counts = {}

            # Fill NaN values with backup columns in priority order
            for backup_col in existing_columns[1:]:
                mask = (
                    df[unified_col_name].isna()
                    & pd.to_numeric(df[backup_col], errors="coerce").notna()
                )
                filled_count = mask.sum()
                if filled_count > 0:
                    df.loc[mask, unified_col_name] = pd.to_numeric(
                        df.loc[mask, backup_col], errors="coerce"
                    )
                    fallback_counts[backup_col] = filled_count

            final_null_count = df[unified_col_name].isna().sum()
            total_recovered = initial_null_count - final_null_count

            if total_recovered > 0:
                logger.info(
                    f"  ✅ {csv_name}: '{parameter_name}' - recovered {total_recovered} values via fallback"
                )
                for col, count in fallback_counts.items():
                    logger.debug(f"      → '{col}': {count} values")
            else:
                logger.debug(
                    f"  {csv_name}: '{parameter_name}' - no fallback needed (primary complete)"
                )


def apply_empirical_calculations(
    df_copy: pd.DataFrame,
    parameter_name: str,
    csv_name: str,
    column_name: str,
    f1_factor_config: Optional[Dict[str, float]] = None,
    geology_code_descriptions: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    """
    Apply empirical calculations for derived parameters with formation-specific factors.

    Args:
        df_copy: DataFrame copy to modify
        parameter_name: Target parameter being analyzed
        csv_name: Source CSV file name
        column_name: Source column name
        f1_factor_config: Dictionary mapping formation codes to f1 factors
        geology_code_descriptions: Geology code descriptions from CONFIG

    Returns:
        DataFrame with empirical calculations applied
    """
    # SPT N-value to Undrained Shear Strength conversion
    if (
        parameter_name == "UndrainedShearStrength"
        and csv_name == "SPT by Geology.csv"
        and column_name == "N"
        and f1_factor_config is not None
    ):
        logger.info(
            f"🔄 Applying formation-specific SPT cu calculations for {csv_name}"
        )

        if "Geological_Strata" not in df_copy.columns:
            logger.warning(
                "⚠️ No Geological_Strata column found - using default f1 factor"
            )
            default_f1 = f1_factor_config.get("default_f1_factor", 4.5)
            df_copy[column_name] = (
                pd.to_numeric(df_copy[column_name], errors="coerce") * default_f1
            )
            return df_copy

        calculated_column = f"{column_name}_calculated_cu"
        default_f1 = f1_factor_config.get("default_f1_factor", 4.5)

        df_copy[calculated_column] = (
            pd.to_numeric(df_copy[column_name], errors="coerce") * default_f1
        )

        formation_counts = {"default": {"f1_factor": default_f1, "conversions": 0}}
        total_conversions = df_copy[calculated_column].notna().sum()

        for formation_code, f1_factor in f1_factor_config.items():
            if formation_code == "default_f1_factor":
                continue

            formation_mask = pd.Series([False] * len(df_copy), index=df_copy.index)

            if "_undefined" in formation_code:
                base_formation = formation_code.replace("_undefined", "")
                if (
                    geology_code_descriptions
                    and base_formation in geology_code_descriptions
                ):
                    target_description = f"{geology_code_descriptions[base_formation]} (Undefined Classification)"
                    undefined_match = df_copy["Geological_Strata"] == target_description
                    formation_mask = formation_mask | undefined_match
            elif formation_code in geology_code_descriptions:
                target_description = geology_code_descriptions[formation_code]
                exact_match = df_copy["Geological_Strata"] == target_description
                formation_mask = formation_mask | exact_match

            if formation_mask.any():
                valid_n_mask = pd.to_numeric(
                    df_copy[column_name], errors="coerce"
                ).notna()
                calculation_mask = formation_mask & valid_n_mask

                df_copy.loc[calculation_mask, calculated_column] = (
                    pd.to_numeric(
                        df_copy.loc[calculation_mask, column_name], errors="coerce"
                    )
                    * f1_factor
                )

                conversions = calculation_mask.sum()
                formation_counts[formation_code] = {
                    "f1_factor": f1_factor,
                    "conversions": conversions,
                }

                logger.info(
                    f"  ✅ {formation_code} (f1={f1_factor}): {conversions} conversions"
                )

        identified_samples = sum(
            info["conversions"]
            for code, info in formation_counts.items()
            if code != "default"
        )
        default_samples = total_conversions - identified_samples
        formation_counts["default"]["conversions"] = default_samples

        if default_samples > 0:
            logger.info(
                f"  ⚠️ Unidentified formations (f1={default_f1}): {default_samples} conversions"
            )

        df_copy[column_name] = df_copy[calculated_column]
        df_copy.drop(columns=[calculated_column], inplace=True)

        logger.info(
            f"✅ Formation-specific SPT calculation completed: {total_conversions} total conversions"
        )

    elif (
        parameter_name == "UndrainedShearStrength"
        and csv_name == "SPT by Geology.csv"
        and column_name == "N"
        and f1_factor_config is None
    ):
        logger.info(
            f"🔄 Applying legacy empirical calculation: cu = f1 × N for {csv_name}"
        )
        calculated_column = f"{column_name}_calculated_cu"
        df_copy[calculated_column] = (
            pd.to_numeric(df_copy[column_name], errors="coerce") * 4.5
        )
        valid_conversions = df_copy[calculated_column].notna().sum()
        total_n_values = df_copy[column_name].notna().sum()
        logger.info(
            f"✅ Legacy empirical calculation completed: {valid_conversions}/{total_n_values} N-values converted to cu"
        )
        df_copy[column_name] = df_copy[calculated_column]
        df_copy.drop(columns=[calculated_column], inplace=True)

    return df_copy


def apply_value_transformation(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    parameter_mappings: pd.DataFrame,
    transformation_config: Dict[str, Any],
    geology_code_descriptions: Dict[str, str],
) -> Dict[str, Dict[str, pd.DataFrame]]:
    """
    Apply value transformation to parameter values by formation.

    Transforms values in-place using formation-specific multipliers and unit conversion.
    Used for derived parameters like Eu (Undrained Stiffness) from Cu (Undrained Shear Strength).

    Formula: transformed_value = (original_value × multiplier) / unit_divisor

    Args:
        formation_groups: Nested dict {formation_name: {csv_name: DataFrame}}
        parameter_mappings: DataFrame with csv_file and column_name mappings
        transformation_config: Configuration dict with:
            - multiplier_by_formation: Dict mapping formation codes to multipliers
            - unit_divisor: Divisor for unit conversion (e.g., 1000 for kPa→MPa)
            - applicable_formations: Optional list of formations to transform (all if None)
        geology_code_descriptions: Mapping of geology codes to descriptions

    Returns:
        Modified formation_groups with transformed values (also modifies in-place)
    """
    multipliers = transformation_config.get("multiplier_by_formation", {})
    unit_divisor = transformation_config.get("unit_divisor", 1)
    applicable_formations = transformation_config.get("applicable_formations")

    # Build reverse lookup: description → formation code
    description_to_code = {v: k for k, v in geology_code_descriptions.items()}

    total_transformed = 0
    formations_transformed = []

    for formation_name, csv_dict in formation_groups.items():
        # Check if this formation should be transformed
        if applicable_formations:
            # Get the formation code from the description
            formation_code = description_to_code.get(formation_name)
            if formation_code not in applicable_formations:
                logger.debug(
                    f"  ⏭️ Skipping {formation_name} - not in applicable formations"
                )
                continue

        # Get formation-specific multiplier
        formation_code = description_to_code.get(formation_name)
        multiplier = multipliers.get(formation_code)

        if multiplier is None:
            # Try with default multiplier
            multiplier = multipliers.get("default")
            if multiplier is None:
                logger.debug(f"  ⏭️ Skipping {formation_name} - no multiplier defined")
                continue

        formations_transformed.append(formation_name)

        for csv_name, df in csv_dict.items():
            # Find the parameter column for this CSV
            matching_row = parameter_mappings[
                parameter_mappings["csv_file"].str.contains(
                    csv_name.replace(".csv", ""), case=False
                )
            ]
            if matching_row.empty:
                continue

            column_name = matching_row.iloc[0]["column_name"]
            if column_name not in df.columns:
                continue

            # Apply transformation: (value × multiplier) / unit_divisor
            original_values = pd.to_numeric(df[column_name], errors="coerce")
            valid_mask = original_values.notna()
            transformed_count = valid_mask.sum()

            df[column_name] = (original_values * multiplier) / unit_divisor
            total_transformed += transformed_count

        logger.debug(
            f"  ✅ {formation_name}: multiplier={multiplier}, divisor={unit_divisor}"
        )

    if formations_transformed:
        logger.info(
            f"✅ Value transformation applied to {len(formations_transformed)} formations, "
            f"{total_transformed} values transformed"
        )
    else:
        logger.warning("⚠️ No formations matched for value transformation")

    return formation_groups


def load_parameter_data(
    parameter_mappings: pd.DataFrame,
    csv_source_folder: str,
    geology_code_descriptions: Dict[str, str],
    column_variations: Dict[str, List[str]],
    formation_splitting_config: Dict[str, Any],
    filter_enabled: bool,
    exclude_prefixes: List[str],
    case_sensitive: bool,
    empirical_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, pd.DataFrame]:
    """
    Load all CSV files containing the target parameter with geological enhancement.

    Args:
        parameter_mappings: DataFrame from extract_parameter_mappings()
        csv_source_folder: Folder containing CSV files
        geology_code_descriptions: Mapping of geology codes to descriptions
        column_variations: Dictionary of column name variations
        formation_splitting_config: Configuration for formation splitting
        filter_enabled: Whether location filtering is enabled
        exclude_prefixes: List of location prefixes to exclude
        case_sensitive: Whether location filtering is case-sensitive
        empirical_config: Optional configuration for empirical calculations

    Returns:
        Dictionary mapping CSV filenames to enhanced DataFrames
    """
    logger.info(f"🚀 PHASE 2: Loading CSV data for parameter sources")

    # Build list of unique CSV file paths
    unique_csv_files = parameter_mappings["csv_file"].unique()
    csv_files = [
        Path(csv_source_folder) / csv_filename for csv_filename in unique_csv_files
    ]

    logger.info(f"📂 Will load {len(csv_files)} unique CSV files:")
    for csv_file in csv_files:
        logger.info(f"  • {csv_file.name}")

    # Load CSV data
    csv_data_dict = load_csv_data(
        csv_files,
        geology_code_descriptions,
        column_variations,
        formation_splitting_config,
        filter_enabled,
        exclude_prefixes,
        case_sensitive,
    )

    # Apply empirical calculations if configured
    if csv_data_dict and empirical_config:
        logger.info("")
        logger.info(f"🔧 Applying empirical calculations...")

        f1_factor_config = empirical_config.get("spt_to_cu_factors")
        parameter_name = empirical_config.get("parameter_name")

        for _, mapping in parameter_mappings.iterrows():
            csv_name = mapping["csv_file"]
            column_name = mapping["column_name"]

            if csv_name in csv_data_dict:
                df_original = csv_data_dict[csv_name]
                df_modified = apply_empirical_calculations(
                    df_original.copy(),
                    parameter_name,
                    csv_name,
                    column_name,
                    f1_factor_config,
                    geology_code_descriptions,
                )
                csv_data_dict[csv_name] = df_modified

        logger.info(f"✅ Empirical calculations complete")
        logger.info("")

    if not csv_data_dict:
        logger.error("❌ PHASE 2 FAILED: No CSV files successfully loaded")
        return {}

    logger.info("")
    logger.info(f"✅ PHASE 2 COMPLETE: Loaded {len(csv_data_dict)} CSV files")
    for csv_name, df in csv_data_dict.items():
        formations = (
            df["Geological_Strata"].nunique()
            if "Geological_Strata" in df.columns
            else 0
        )
        logger.info(f"  • {csv_name}: {len(df)} rows, {formations} formations")

    return csv_data_dict


# ═══════════════════════════════════════════════════════════════════════════
# 🔄 SECTION 4: FORMATION GROUPING & DATA ORGANIZATION
# ═══════════════════════════════════════════════════════════════════════════


def group_data_by_formation_unified(
    csv_data_dict: Dict[str, pd.DataFrame],
    x_param_mappings: pd.DataFrame,
    y_param_mappings: Optional[pd.DataFrame] = None,
    x_column: Optional[str] = None,
    y_column: str = "Top Depth",
) -> Dict[str, Dict[str, pd.DataFrame]]:
    """
    Unified formation grouping for all plot types (depth and XY).

    Groups data by geological formation from CSV files, supporting both
    depth plots (parameter vs depth) and XY plots (parameter vs parameter).

    For depth plots: y_column = "Top Depth" (default)
    For XY plots: y_column = y_parameter name, y_param_mappings required

    Args:
        csv_data_dict: Dictionary of loaded and enhanced CSV data
        x_param_mappings: X parameter mappings (or single parameter for depth plots)
        y_param_mappings: Y parameter mappings (None for depth plots - uses Top Depth)
        x_column: Explicit X column name. If None, derived from x_param_mappings
        y_column: Y column name. Default "Top Depth" for depth plots.
                  For XY plots, pass the y_parameter column name.

    Returns:
        Nested dictionary: {formation_name: {csv_name: dataframe}}
        Only includes rows where BOTH X and Y columns have valid values.
    """
    is_dual_param = y_param_mappings is not None and y_column != "Top Depth"
    mode_label = "dual-param" if is_dual_param else "depth"

    logger.info(
        f"🚀 PHASE 3: Grouping data by geological formation ({mode_label} mode)"
    )

    formation_dict = defaultdict(lambda: defaultdict(pd.DataFrame))

    # For dual-param, find CSVs that have BOTH parameters
    if is_dual_param:
        x_csvs = set(x_param_mappings["csv_file"].unique())
        y_csvs = set(y_param_mappings["csv_file"].unique())
        common_csvs = x_csvs & y_csvs
    else:
        common_csvs = None  # Use all CSVs for depth plots

    for csv_name, df in csv_data_dict.items():
        # For dual-param mode, skip CSVs not in common set
        if common_csvs is not None and csv_name not in common_csvs:
            logger.debug(f"⏭️ {csv_name}: Skipping - not in common CSV set")
            continue

        logger.debug(f"  Processing {csv_name}...")

        # === DETERMINE X COLUMN ===
        if x_column is not None:
            resolved_x_column = x_column
        else:
            # Get from mappings
            x_mapping = x_param_mappings[x_param_mappings["csv_file"] == csv_name]
            if x_mapping.empty:
                logger.warning(f"⚠️ {csv_name}: No X parameter mapping found, skipping")
                continue
            x_param_name = x_mapping["parameter"].iloc[0]
            x_unified_col = f"{x_param_name}_unified"
            x_original_col = x_mapping["column_name"].iloc[0]
            resolved_x_column = (
                x_unified_col if x_unified_col in df.columns else x_original_col
            )

        # === DETERMINE Y COLUMN ===
        if y_column == "Top Depth":
            resolved_y_column = "Top Depth"
        elif is_dual_param:
            # Get Y column from mappings
            y_mapping = y_param_mappings[y_param_mappings["csv_file"] == csv_name]
            if y_mapping.empty:
                logger.warning(f"⚠️ {csv_name}: No Y parameter mapping found, skipping")
                continue
            y_param_name = y_mapping["parameter"].iloc[0]
            y_unified_col = f"{y_param_name}_unified"
            y_original_col = y_mapping["column_name"].iloc[0]
            resolved_y_column = (
                y_unified_col if y_unified_col in df.columns else y_original_col
            )
        else:
            resolved_y_column = y_column

        # === VALIDATE COLUMNS ===
        if resolved_x_column not in df.columns:
            logger.warning(
                f"⚠️ {csv_name}: X column '{resolved_x_column}' not found, skipping"
            )
            continue

        if resolved_y_column not in df.columns:
            logger.warning(
                f"⚠️ {csv_name}: Y column '{resolved_y_column}' not found, skipping"
            )
            continue

        if "Geological_Strata" not in df.columns:
            logger.warning(
                f"⚠️ {csv_name}: Geological_Strata column not found, skipping"
            )
            continue

        # Log which columns are being used
        logger.info(
            f"  📊 {csv_name}: Using X='{resolved_x_column}', Y='{resolved_y_column}'"
        )

        formations = df["Geological_Strata"].unique()

        for formation in formations:
            formation_data = df[df["Geological_Strata"] == formation].copy()

            # Drop rows with null values in EITHER X or Y column
            formation_data = formation_data.dropna(
                subset=[resolved_x_column, resolved_y_column]
            )

            # Ensure values are valid numerics
            formation_data = formation_data[
                pd.to_numeric(
                    formation_data[resolved_x_column], errors="coerce"
                ).notna()
                & pd.to_numeric(
                    formation_data[resolved_y_column], errors="coerce"
                ).notna()
            ]

            if not formation_data.empty:
                formation_dict[formation][csv_name] = formation_data
                logger.debug(
                    f"    → {formation}: {len(formation_data)} samples from {csv_name}"
                )

    logger.info("")
    logger.info(f"✅ Grouped data into {len(formation_dict)} formations:")

    for formation, csv_dict in sorted(formation_dict.items()):
        total_samples = sum(len(df) for df in csv_dict.values())
        csv_count = len(csv_dict)
        logger.info(
            f"  • {formation}: {total_samples} samples from {csv_count} CSV source(s)"
        )

    return dict(formation_dict)


def apply_cell_pressure_threshold_filtering(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    x_param_mappings: pd.DataFrame,
    cell_pressure_threshold_config: Dict[str, Any],
) -> None:
    """
    Apply cell pressure threshold filtering to formation groups (in-place).

    Removes data points where the cell pressure (X parameter) exceeds the
    configured maximum threshold. This is used for consolidation plots to
    filter out unrealistic high-pressure data points.

    Args:
        formation_groups: Formation-grouped data (MODIFIED IN-PLACE)
        x_param_mappings: X parameter mappings (contains Cell Pressure column info)
        cell_pressure_threshold_config: Configuration with 'enabled' and 'max_value' keys
    """
    if not cell_pressure_threshold_config.get("enabled", False):
        return

    max_pressure = cell_pressure_threshold_config.get("max_value", 1500)
    total_removed = 0

    logger.info(
        f"🔍 Applying cell pressure threshold filter (max: {max_pressure} kPa)..."
    )

    for formation_name, csv_dict in formation_groups.items():
        for csv_name, df in list(csv_dict.items()):
            # Find cell pressure column for this CSV
            csv_x_mappings = x_param_mappings[x_param_mappings["csv_file"] == csv_name]
            if csv_x_mappings.empty:
                continue

            cell_pressure_col = csv_x_mappings["column_name"].iloc[0]

            # Check for unified column first
            unified_col = f"{csv_x_mappings['parameter'].iloc[0]}_unified"
            if unified_col in df.columns:
                cell_pressure_col = unified_col

            if cell_pressure_col not in df.columns:
                continue

            initial_count = len(df)
            filtered_df = df[df[cell_pressure_col] <= max_pressure].copy()
            removed_count = initial_count - len(filtered_df)

            if removed_count > 0:
                csv_dict[csv_name] = filtered_df
                total_removed += removed_count
                logger.debug(
                    f"  → {formation_name} ({csv_name}): Removed {removed_count} samples "
                    f"with {cell_pressure_col} > {max_pressure} kPa"
                )

    if total_removed > 0:
        logger.info(
            f"✅ Cell pressure threshold filtering: Removed {total_removed} samples total"
        )


def apply_formation_parameter_threshold_filtering(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    parameter_mappings: pd.DataFrame,
    formation_threshold_config: Dict[str, Any],
) -> None:
    """
    Apply formation-specific parameter threshold filtering to formation groups (in-place).

    Removes data points that exceed formation-specific parameter thresholds.
    For example, filtering KC_UW to remove cv values > 5.0 m²/year.

    Args:
        formation_groups: Formation-grouped data (MODIFIED IN-PLACE)
        parameter_mappings: Parameter mappings (for finding column names)
        formation_threshold_config: Configuration with 'enabled' and 'thresholds' keys

    Example config:
        {
            "enabled": True,
            "thresholds": {
                "Kimmeridge Clay Formation (Unweathered)": {
                    "CoefficientConsolidation": {"max_value": 5.0}
                }
            }
        }
    """
    if not formation_threshold_config.get("enabled", False):
        return

    thresholds = formation_threshold_config.get("thresholds", {})
    if not thresholds:
        return

    total_removed = 0
    logger.info(f"🔍 Applying formation-specific parameter threshold filters...")

    for formation_name, csv_dict in formation_groups.items():
        if formation_name not in thresholds:
            continue

        formation_thresholds = thresholds[formation_name]

        for csv_name, df in list(csv_dict.items()):
            for param_name, param_threshold in formation_thresholds.items():
                # Find the column for this parameter
                param_mapping = parameter_mappings[
                    (parameter_mappings["csv_file"] == csv_name)
                    & (parameter_mappings["parameter"] == param_name)
                ]

                if param_mapping.empty:
                    continue

                param_col = param_mapping["column_name"].iloc[0]

                # Check for unified column first (in case of row-level fallback)
                unified_col = f"{param_name}_unified"
                if unified_col in df.columns:
                    param_col = unified_col

                if param_col not in df.columns:
                    continue

                initial_count = len(df)

                # Apply max_value threshold
                if "max_value" in param_threshold:
                    max_val = param_threshold["max_value"]
                    df = df[df[param_col] <= max_val].copy()

                # Apply min_value threshold
                if "min_value" in param_threshold:
                    min_val = param_threshold["min_value"]
                    df = df[df[param_col] >= min_val].copy()

                removed_count = initial_count - len(df)

                if removed_count > 0:
                    csv_dict[csv_name] = df
                    total_removed += removed_count
                    logger.info(
                        f"  → {formation_name} ({csv_name}): Removed {removed_count} samples "
                        f"with {param_name} threshold"
                    )

    if total_removed > 0:
        logger.info(
            f"✅ Formation parameter threshold filtering: Removed {total_removed} samples total"
        )


def apply_test_type_threshold_filtering(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    parameter_mappings: pd.DataFrame,
    csv_source_settings: Dict[str, Any],
) -> None:
    """
    Apply test-type-specific threshold filtering to formation groups (in-place).

    ═══════════════════════════════════════════════════════════════════════════════
    ⚠️ CRITICAL ARCHITECTURAL REQUIREMENT - READ BEFORE MODIFYING ⚠️
    ═══════════════════════════════════════════════════════════════════════════════

    This function is part of the SCRIPT-SPECIFIC FILTERING phase that MUST execute
    BEFORE prepare_unified_data_objects() is called. This ensures that:

    1. COUNTS MATCH PLOTTED DATA: Investigation counts and test type counts are
       calculated from unified data objects. If filtering happens AFTER unification,
       counts will include filtered-out data points, causing user confusion.

    2. SINGLE SOURCE OF TRUTH: Unified data objects should be the single source of
       truth for all plotting and count calculations. ALL data filtering must be
       complete before these objects are created.

    3. NO DUPLICATE FILTERING: By filtering here, we avoid duplicating filter logic
       in plotly_utils.py and matplotlib_utils.py. One filter location = one bug fix.

    ═══════════════════════════════════════════════════════════════════════════════
    ❌ ANTI-PATTERN - DO NOT DO THIS:
    ═══════════════════════════════════════════════════════════════════════════════

    # WRONG: Filtering inside plotting loop - counts won't match plotted data!
    # for inv_id in investigations:
    #     inv_data = df[df["Investigation"] == inv_id]
    #     if max_kpa := settings.get("max_kpa"):
    #         inv_data = inv_data[inv_data[param_col] <= max_kpa]  # ❌ TOO LATE!
    #     # Counts were already calculated before this loop...

    ═══════════════════════════════════════════════════════════════════════════════
    ✅ CORRECT PATTERN - THIS FUNCTION IMPLEMENTS:
    ═══════════════════════════════════════════════════════════════════════════════

    # RIGHT: Filter in orchestrator Phase 3 BEFORE prepare_unified_data_objects()
    # formation_groups modified in-place → clean data flows to unification
    # → counts calculated from clean data → counts match plotted data ✅

    ═══════════════════════════════════════════════════════════════════════════════

    Removes data points where test-type-specific thresholds are exceeded.
    Example: Vane Tests with max_kpa=129 filters unreliable hand vane readings.

    Args:
        formation_groups: Formation-grouped data (MODIFIED IN-PLACE)
        parameter_mappings: Parameter mappings (for finding column names)
        csv_source_settings: CSV source settings containing test-type thresholds
            Example: {"Vane Tests by Geology.csv": {"max_kpa": 129, ...}}

    Supported threshold keys in csv_source_settings:
        - max_kpa: Maximum parameter value (data points > max_kpa are removed)
        - min_kpa: Minimum parameter value (data points < min_kpa are removed)
    """
    if not csv_source_settings:
        return

    total_removed = 0
    logger.info("🔍 Applying test-type-specific threshold filters...")

    for formation_name, csv_dict in formation_groups.items():
        for csv_name, df in list(csv_dict.items()):
            # Get settings for this CSV/test type
            settings = csv_source_settings.get(csv_name, {})

            # Check for max_kpa threshold
            max_kpa = settings.get("max_kpa")
            min_kpa = settings.get("min_kpa")

            if max_kpa is None and min_kpa is None:
                continue

            # Find the parameter column for this CSV
            csv_mappings = parameter_mappings[
                parameter_mappings["csv_file"] == csv_name
            ]
            if csv_mappings.empty:
                continue

            param_col = csv_mappings["column_name"].iloc[0]
            param_name = csv_mappings["parameter"].iloc[0]

            # Check for unified column first (in case of row-level fallback)
            unified_col = f"{param_name}_unified"
            if unified_col in df.columns:
                param_col = unified_col

            if param_col not in df.columns:
                continue

            initial_count = len(df)

            # Apply max_kpa threshold
            if max_kpa is not None:
                df = df[df[param_col] <= max_kpa].copy()

            # Apply min_kpa threshold
            if min_kpa is not None:
                df = df[df[param_col] >= min_kpa].copy()

            removed_count = initial_count - len(df)

            if removed_count > 0:
                csv_dict[csv_name] = df
                total_removed += removed_count
                threshold_desc = []
                if max_kpa is not None:
                    threshold_desc.append(f"max_kpa={max_kpa}")
                if min_kpa is not None:
                    threshold_desc.append(f"min_kpa={min_kpa}")
                logger.info(
                    f"  → {formation_name} ({csv_name}): Removed {removed_count} samples "
                    f"with {', '.join(threshold_desc)} threshold"
                )

    if total_removed > 0:
        logger.info(
            f"✅ Test-type threshold filtering: Removed {total_removed} samples total"
        )


# ═══════════════════════════════════════════════════════════════════════════
# 🔍 SECTION 5: OUTLIER DETECTION
# ═══════════════════════════════════════════════════════════════════════════


def detect_outliers_standard_iqr(
    values_clean: pd.Series,
    iqr_multiplier: float = 1.5,
    q1_boundary: float = 0.25,
    q3_boundary: float = 0.75,
) -> Tuple[pd.Series, pd.Series, float, float]:
    """
    Detect outliers using the standard Tukey boxplot method.

    Args:
        values_clean: Clean numerical data (no NaNs)
        iqr_multiplier: Multiplier for IQR (default 1.5)
        q1_boundary: First quartile boundary (default 0.25)
        q3_boundary: Third quartile boundary (default 0.75)

    Returns:
        Tuple of (mild_outliers, extreme_outliers, lower_bound, upper_bound)
    """
    if values_clean.empty or values_clean.isna().all():
        empty_series = pd.Series([], dtype=float)
        return empty_series, empty_series, float("nan"), float("nan")

    q1 = values_clean.quantile(q1_boundary)
    q3 = values_clean.quantile(q3_boundary)
    iqr = q3 - q1

    if iqr == 0:
        logger.debug(
            f"Zero IQR detected for dataset with {len(values_clean)} values - no outliers identified"
        )
        empty_series = pd.Series([], dtype=float)
        return empty_series, empty_series, float("nan"), float("nan")

    lower_fence = q1 - iqr_multiplier * iqr
    upper_fence = q3 + iqr_multiplier * iqr

    outlier_mask = (values_clean < lower_fence) | (values_clean > upper_fence)
    outliers = values_clean[outlier_mask]

    logger.debug(
        f"Standard IQR outlier detection: {len(outliers)} outliers from {len(values_clean)} samples "
        f"(bounds: [{lower_fence:.3f}, {upper_fence:.3f}])"
    )

    return (
        outliers,
        pd.Series([], dtype=float),
        float(lower_fence),
        float(upper_fence),
    )


def detect_outliers_per_csv(
    csv_dict: Dict[str, pd.DataFrame],
    parameter_mappings: pd.DataFrame,
    formation_name: str,
    enabled: bool,
    min_samples: int,
    iqr_multiplier: float,
    q1_boundary: float,
    q3_boundary: float,
) -> Dict[str, Dict[str, Any]]:
    """
    Detect outliers per test type (CSV) within a formation.

    Args:
        csv_dict: Dictionary of {csv_name: dataframe} for one formation
        parameter_mappings: Parameter mapping dataframe
        formation_name: Name of the formation being processed
        enabled: Whether outlier detection is enabled
        min_samples: Minimum samples required for outlier detection
        iqr_multiplier: IQR multiplier for outlier bounds
        q1_boundary: First quartile boundary
        q3_boundary: Third quartile boundary

    Returns:
        Dictionary of outlier results per CSV
    """
    outlier_results = {}

    if not enabled:
        logger.info(f"📊 Outlier detection disabled for {formation_name}")
        return outlier_results

    logger.info(f"🔍 Detecting outliers per test type for {formation_name}...")

    for csv_name, df in csv_dict.items():
        param_row = parameter_mappings[parameter_mappings["csv_file"] == csv_name]
        if param_row.empty:
            continue

        param_column = param_row.iloc[0]["column_name"]

        # Check for unified column first (in case of row-level fallback)
        param_name = param_row.iloc[0]["parameter"]
        unified_col = f"{param_name}_unified"
        if unified_col in df.columns:
            param_column = unified_col

        if param_column not in df.columns:
            logger.warning(
                f"⚠️ Parameter column '{param_column}' not found in {csv_name}"
            )
            continue

        values_series = df[param_column]
        values_clean = pd.to_numeric(values_series, errors="coerce").dropna()

        if len(values_clean) < min_samples:
            logger.debug(
                f"  ⏭️ {csv_name}: {len(values_clean)} samples < {min_samples} minimum - skipping"
            )
            outlier_results[csv_name] = {
                "outlier_indices": set(),
                "outlier_count": 0,
                "total_samples": len(values_clean),
            }
            continue

        mild_outliers, _, lower_bound, upper_bound = detect_outliers_standard_iqr(
            values_clean, iqr_multiplier, q1_boundary, q3_boundary
        )

        outlier_indices = set(mild_outliers.index.tolist())
        outlier_count = len(outlier_indices)
        total_samples = len(values_clean)

        outlier_results[csv_name] = {
            "outlier_indices": outlier_indices,
            "outlier_count": outlier_count,
            "total_samples": total_samples,
            "lower_bound": lower_bound,
            "upper_bound": upper_bound,
        }

        logger.info(
            f"  ✅ {csv_name}: {outlier_count}/{total_samples} outliers detected "
            f"(bounds: [{lower_bound:.2f}, {upper_bound:.2f}])"
        )

    return outlier_results


def filter_outliers_from_formations(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    parameter_mappings: pd.DataFrame,
    enabled: bool,
    min_samples: int,
    iqr_multiplier: float,
    q1_boundary: float,
    q3_boundary: float,
) -> None:
    """
    Apply outlier filtering to all formations (modifies in-place).

    Args:
        formation_groups: Dict mapping formations to CSV data dicts
        parameter_mappings: Parameter mapping DataFrame
        enabled: Whether outlier detection is enabled
        min_samples: Minimum samples for outlier detection
        iqr_multiplier: IQR multiplier for bounds
        q1_boundary: First quartile boundary
        q3_boundary: Third quartile boundary
    """
    if not enabled:
        logger.info("⏭️ Outlier filtering disabled - skipping Phase 3b")
        return

    logger.info(
        f"🚀 PHASE 3b: Filtering outliers for {len(formation_groups)} formations"
    )

    for formation_name in sorted(formation_groups.keys()):
        csv_dict = formation_groups[formation_name]

        outlier_results = detect_outliers_per_csv(
            csv_dict,
            parameter_mappings,
            formation_name,
            enabled,
            min_samples,
            iqr_multiplier,
            q1_boundary,
            q3_boundary,
        )

        for csv_name, df in csv_dict.items():
            if csv_name in outlier_results:
                outlier_indices = outlier_results[csv_name]["outlier_indices"]
                df["is_outlier"] = df.index.isin(outlier_indices)
            else:
                df["is_outlier"] = False

    logger.info("✅ Outlier filtering complete - 'is_outlier' column added to all data")


# ═══════════════════════════════════════════════════════════════════════════
# 🎨 SECTION 6: INVESTIGATION MAPPING & UNIFIED DATA OBJECTS
# ═══════════════════════════════════════════════════════════════════════════


def load_location_details(
    csv_folder: str,
    location_csv_name: str,
    location_id_variants: List[str],
    investigation_variants: List[str],
) -> pd.DataFrame:
    """
    Load Location Details CSV for per-investigation plotting.

    Args:
        csv_folder: Path to folder containing Location Details CSV
        location_csv_name: Name of the Location Details CSV file
        location_id_variants: List of possible Location ID column names
        investigation_variants: List of possible Investigation column names

    Returns:
        DataFrame with standardized 'Location ID' and 'Investigation' columns
    """
    csv_folder_path = Path(csv_folder)
    location_csv = csv_folder_path / location_csv_name

    try:
        df = pd.read_csv(location_csv, encoding="utf-8-sig")
        df.columns = df.columns.str.strip()

        location_id_col = None
        for variant in location_id_variants:
            if variant in df.columns:
                location_id_col = variant
                break

        investigation_col = None
        for variant in investigation_variants:
            if variant in df.columns:
                investigation_col = variant
                break

        if not location_id_col or not investigation_col:
            logger.warning(
                f"⚠️ Required columns not found in {location_csv_name}. "
                f"Expected: {location_id_variants} and {investigation_variants}"
            )
            return pd.DataFrame()

        result_df = df[[location_id_col, investigation_col]].copy()
        result_df.columns = ["Location ID", "Investigation"]
        result_df["Location ID"] = result_df["Location ID"].astype(str).str.strip()
        result_df["Investigation"] = result_df["Investigation"].astype(str).str.strip()

        return result_df

    except FileNotFoundError:
        logger.warning(f"⚠️ Location Details CSV not found: {location_csv}")
        return pd.DataFrame()
    except Exception as e:
        logger.warning(f"⚠️ Failed to load location details: {str(e)}")
        return pd.DataFrame()


def add_investigation_column(
    csv_data: pd.DataFrame,
    location_details: pd.DataFrame,
    fallback_name: str,
) -> pd.DataFrame:
    """
    Add Investigation column to CSV data by joining with Location Details.

    Args:
        csv_data: Data from single CSV source (must have 'Location ID' column)
        location_details: DataFrame with 'Location ID' and 'Investigation' columns
        fallback_name: Default investigation name for missing/unmatched IDs

    Returns:
        DataFrame with new 'Investigation' column added
    """
    if location_details.empty or "Location ID" not in csv_data.columns:
        result = csv_data.copy()
        result["Investigation"] = fallback_name
        return result

    result = csv_data.merge(
        location_details,
        on="Location ID",
        how="left",
    )
    result["Investigation"] = result["Investigation"].fillna(fallback_name)

    return result


def create_investigation_color_mapping(
    investigations: List[str],
    color_palette: List[str],
) -> Dict[str, str]:
    """
    Create color mapping for investigations using rotating color palette.

    Args:
        investigations: List of unique investigation names
        color_palette: List of hex color codes

    Returns:
        Dictionary mapping investigation names to hex color codes
    """
    investigation_colors = {}
    for idx, investigation in enumerate(sorted(investigations)):
        color_idx = idx % len(color_palette)
        investigation_colors[investigation] = color_palette[color_idx]
    return investigation_colors


def prepare_unified_data_objects(
    formation_groups: Dict[str, Dict[str, pd.DataFrame]],
    location_details: pd.DataFrame,
    fallback_investigation_name: str,
    csv_source_settings: Dict[str, Dict[str, Any]],
    default_source_settings: Dict[str, Any],
) -> Tuple[Dict[str, Dict[str, pd.DataFrame]], Dict[str, Dict[str, pd.DataFrame]]]:
    """
    Create unified data objects after all filtering is complete.

    Args:
        formation_groups: Formation data after Phase 3c
        location_details: DataFrame with Location ID -> Investigation mapping
        fallback_investigation_name: Default name for missing investigations
        csv_source_settings: Settings for each CSV source
        default_source_settings: Default settings for unmapped sources

    Returns:
        Tuple of (plotly_unified_data, manually_modified_data)
    """
    logger.info("🚀 Creating unified data objects for downstream outputs...")

    plotly_unified_data = {}
    manually_modified_data = {}

    for formation_name, csv_dict in formation_groups.items():
        plotly_unified_data[formation_name] = {}
        manually_modified_data[formation_name] = {}

        for csv_name, df in csv_dict.items():
            csv_settings = csv_source_settings.get(csv_name, default_source_settings)
            if not csv_settings.get("plotted", True):
                continue

            df_with_inv = add_investigation_column(
                df.copy(), location_details, fallback_investigation_name
            )

            if "is_outlier" not in df_with_inv.columns:
                df_with_inv["is_outlier"] = False
            if "is_manual_outlier" not in df_with_inv.columns:
                df_with_inv["is_manual_outlier"] = False
            if "is_manual_addition" not in df_with_inv.columns:
                df_with_inv["is_manual_addition"] = False

            plotly_unified_data[formation_name][csv_name] = df_with_inv.copy()

            mask = ~df_with_inv["is_manual_outlier"] | df_with_inv["is_manual_addition"]
            manually_modified_data[formation_name][csv_name] = df_with_inv[mask].copy()

    plotly_formations = len([f for f, d in plotly_unified_data.items() if d])
    manual_formations = len([f for f, d in manually_modified_data.items() if d])

    plotly_total = sum(
        sum(len(df) for df in csv_dict.values())
        for csv_dict in plotly_unified_data.values()
    )
    manual_total = sum(
        sum(len(df) for df in csv_dict.values())
        for csv_dict in manually_modified_data.values()
    )

    logger.info(f"✅ Created unified data objects:")
    logger.info(
        f"   📊 plotly_unified_data: {plotly_formations} formations, {plotly_total} total points"
    )
    logger.info(
        f"   📊 manually_modified_data: {manual_formations} formations, {manual_total} total points"
    )
    logger.info(f"   📊 Difference (excluded): {plotly_total - manual_total} points")

    return plotly_unified_data, manually_modified_data


# ═══════════════════════════════════════════════════════════════════════════
# 📤 SECTION 7: EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════


def export_highlighted_excel(
    df: pd.DataFrame,
    output_path: Path,
    outlier_indices: set,
    highlight_color: str = "FFFF0000",
    freeze_header: bool = True,
) -> None:
    """
    Export DataFrame to Excel with outlier rows highlighted in red.

    Args:
        df: DataFrame to export
        output_path: Path to save Excel file
        outlier_indices: Set of row indices to highlight as outliers
        highlight_color: Fill color in ARGB format (default: red)
        freeze_header: Whether to freeze the first row
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write header row
    for col_idx, column_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(column_name))

    red_fill = PatternFill(
        start_color=highlight_color, end_color=highlight_color, fill_type="solid"
    )

    # Write data rows and apply highlighting
    for row_idx, (df_idx, row) in enumerate(df.iterrows(), start=2):
        is_outlier = df_idx in outlier_indices

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if pd.isna(value):
                cell.value = None
            elif isinstance(value, (int, float, np.integer, np.floating)):
                cell.value = float(value) if not pd.isna(value) else None
            else:
                cell.value = str(value)

            if is_outlier:
                cell.fill = red_fill

    if freeze_header:
        ws.freeze_panes = ws["A2"]

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    logger.info(f"  📄 Exported: {output_path.name}")
